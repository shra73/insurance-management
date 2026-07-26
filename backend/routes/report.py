import io
from datetime import datetime
from flask import Blueprint, send_file, jsonify
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)
from models.customer import Customer
from utils.decorators import roles_required
import io as _io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

report_bp = Blueprint("report", __name__, url_prefix="/api/reports")


def _add_header_footer(canvas, doc):
    """Draws a running header and footer (with page number) on every page."""
    canvas.saveState()

    # Header
    canvas.setFont("Helvetica-Bold", 9)
    canvas.setFillColor(colors.HexColor("#555555"))
    canvas.drawString(2 * cm, A4[1] - 1.3 * cm, "Insurance Management Platform")
    canvas.line(2 * cm, A4[1] - 1.4 * cm, A4[0] - 2 * cm, A4[1] - 1.4 * cm)

    # Footer with page number
    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(A4[0] / 2, 1 * cm, f"Page {doc.page}")

    canvas.restoreState()


@report_bp.route("/customers/pdf", methods=["GET"])
@roles_required("ADMIN", "AGENT")
def customer_report_pdf():
    try:
        # Fetch only the columns needed for the report, ordered by name.
        # No relationships (e.g. policies) are loaded — keeps the query
        # lightweight and avoids pulling unrelated data into memory.
        customers = Customer.query.order_by(Customer.name.asc()).all()

        # In-memory byte buffer — the PDF is built entirely in RAM and
        # streamed back to the client; nothing is ever written to a
        # temporary file on disk.
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=2.2 * cm,
            bottomMargin=1.8 * cm,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            title="Customer Report"
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Title"], fontSize=20,
            alignment=TA_CENTER, textColor=colors.HexColor("#1a2b4c")
        )
        subtitle_style = ParagraphStyle(
            "ReportSubtitle", parent=styles["Normal"], fontSize=13,
            alignment=TA_CENTER, textColor=colors.HexColor("#444444"), spaceAfter=4
        )
        date_style = ParagraphStyle(
            "ReportDate", parent=styles["Normal"], fontSize=9,
            alignment=TA_CENTER, textColor=colors.HexColor("#777777")
        )

        story = []
        story.append(Paragraph("Insurance Management Platform", title_style))
        story.append(Paragraph("Customer Report", subtitle_style))
        story.append(Paragraph(
            f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
            date_style
        ))
        story.append(Spacer(1, 16))

        # Table header + one row per customer
        table_data = [["ID", "Customer Name", "Email", "Phone", "Address", "DOB"]]
        for c in customers:
            table_data.append([
                str(c.id),
                c.name,
                c.email,
                c.phone,
                c.address,
                c.dob.isoformat() if c.dob else ""
            ])

        table = Table(
            table_data,
            colWidths=[1.5 * cm, 3.5 * cm, 4.5 * cm, 2.8 * cm, 4 * cm, 2.2 * cm],
            repeatRows=1  # header row repeats on every page for multi-page tables
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2b4c")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)

        story.append(Spacer(1, 14))
        story.append(Paragraph(
            f"<b>Total Customers:</b> {len(customers)}",
            ParagraphStyle("TotalStyle", parent=styles["Normal"], fontSize=11)
        ))

        # Header/footer callback runs on every page, including subsequent
        # pages generated automatically if the table overflows one page.
        doc.build(story, onFirstPage=_add_header_footer, onLaterPages=_add_header_footer)

        buffer.seek(0)

        return send_file(
            buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name="customer_report.pdf"
        )

    except Exception:
        return jsonify({
            "error": "An unexpected error occurred while generating the customer report"
        }), 500



@report_bp.route("/customers/excel", methods=["GET"])
@roles_required("ADMIN", "AGENT")
def customer_report_excel():
    try:
        customers = Customer.query.order_by(Customer.name.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"

        headers = ["ID", "Customer Name", "Email", "Phone", "Address", "Date Of Birth", "Created At"]
        ws.append(headers)

        # --- Header styling ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A2B4C", end_color="1A2B4C", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # --- Thin border applied to every cell (header + data) ---
        thin_side = Side(style="thin", color="CCCCCC")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # --- Data rows ---
        for customer in customers:
            ws.append([
                customer.id,
                customer.name,
                customer.email,
                customer.phone,
                customer.address,
                customer.dob.isoformat() if customer.dob else "",
                customer.created_at.strftime("%Y-%m-%d %H:%M:%S") if customer.created_at else ""
            ])

        # Apply borders to every populated cell (header row + all data rows)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # --- Auto-adjust column widths based on the longest value in each column ---
        for col_num, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_num)
            max_length = len(header)
            for row_num in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = max_length + 4

        # --- Freeze the header row so it stays visible while scrolling ---
        ws.freeze_panes = "A2"

        # --- Save to an in-memory buffer, never to a temp file on disk ---
        buffer = _io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="customer_report.xlsx"
        )

    except Exception:
        return jsonify({
            "error": "An unexpected error occurred while generating the customer Excel report"
        }), 500
    
    