import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _pdf_header_footer(canvas, doc):
    """Shared header/footer drawn on every page of every PDF report."""
    canvas.saveState()
    canvas.setFont("Helvetica-Bold", 9)
    canvas.setFillColor(colors.HexColor("#555555"))
    canvas.drawString(2 * cm, A4[1] - 1.3 * cm, "Insurance Management Platform")
    canvas.line(2 * cm, A4[1] - 1.4 * cm, A4[0] - 2 * cm, A4[1] - 1.4 * cm)
    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(A4[0] / 2, 1 * cm, f"Page {doc.page}")
    canvas.restoreState()


def build_pdf_report(subtitle, headers, rows, total_label, col_widths):
    """
    Builds a professional A4 PDF report and returns an in-memory BytesIO buffer.

    subtitle:   report-specific title, e.g. "Policy Report"
    headers:    list of column header strings
    rows:       list of lists, each inner list matching the header order
    total_label: e.g. "Total Policies" — shown with len(rows) at the end
    col_widths: list of column widths (in cm) matching headers length
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=2.2 * cm, bottomMargin=1.8 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
        title=subtitle
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontSize=20,
        alignment=TA_CENTER, textColor=colors.HexColor("#1a2b4c")
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"], fontSize=13,
        alignment=TA_CENTER, textColor=colors.HexColor("#444444"), spaceAfter=4
    )
    date_style = ParagraphStyle(
        "ReportDate", parent=styles["Normal"], fontSize=9,
        alignment=TA_CENTER, textColor=colors.HexColor("#777777")
    )

    story = [
        Paragraph("Insurance Management Platform", title_style),
        Paragraph(subtitle, subtitle_style),
        Paragraph(f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", date_style),
        Spacer(1, 16)
    ]

    table_data = [headers] + rows
    table = Table(
        table_data,
        colWidths=[w * cm for w in col_widths],
        repeatRows=1
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2b4c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    story.append(Spacer(1, 14))
    story.append(Paragraph(
        f"<b>{total_label}:</b> {len(rows)}",
        ParagraphStyle("TotalStyle", parent=styles["Normal"], fontSize=11)
    ))

    doc.build(story, onFirstPage=_pdf_header_footer, onLaterPages=_pdf_header_footer)
    buffer.seek(0)
    return buffer


def build_excel_report(sheet_name, headers, rows):
    """
    Builds a formatted Excel workbook and returns an in-memory BytesIO buffer.

    sheet_name: e.g. "Policies"
    headers:    list of column header strings
    rows:       list of lists, each inner list matching the header order
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A2B4C", end_color="1A2B4C", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    thin_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row_values in rows:
        ws.append(row_values)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    for col_num, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_num)
        max_length = len(header)
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = max_length + 4

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer